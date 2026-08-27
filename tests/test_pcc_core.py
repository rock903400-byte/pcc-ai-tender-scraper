# -*- coding: utf-8 -*-
"""
pcc_core 的離線單元測試。

所有測試都以 tests/fixtures/ 下的真實回應存檔為輸入，
完全不連線政府採購網。
"""

from datetime import date, timedelta

import pytest

import pcc_core as core


# ==================== 日期與金額 ====================

class TestDateConversion:
    @pytest.mark.parametrize("ad, roc", [
        ("2026/08/24", "115/08/24"),
        ("2026-01-05", "115/01/05"),
        ("2000/12/31", "89/12/31"),
    ])
    def test_to_roc_date(self, ad, roc):
        assert core.to_roc_date(ad) == roc

    @pytest.mark.parametrize("roc, ad", [
        ("115/08/24", "2026/08/24"),
        ("115/9/3", "2026/09/03"),
        ("89/12/31", "2000/12/31"),
    ])
    def test_to_ad_date(self, roc, ad):
        assert core.to_ad_date(roc) == ad

    def test_to_ad_date_passes_through_western_dates(self):
        """已是西元的日期不可被再加 1911。"""
        assert core.to_ad_date("2026/08/24") == "2026/08/24"

    @pytest.mark.parametrize("bad", ["", "未定", "115/08", "N/A"])
    def test_to_ad_date_survives_bad_input(self, bad):
        assert core.to_ad_date(bad) == bad.strip()

    def test_roc_ad_round_trip(self):
        assert core.to_ad_date(core.to_roc_date("2026/08/24")) == "2026/08/24"


class TestParseAmount:
    @pytest.mark.parametrize("text, expected", [
        ("2,608,500 元", 2608500.0),
        ("21,000,000", 21000000.0),
        ("0 元", 0.0),
        ("", -1.0),
        ("未公開", -1.0),
        (None, -1.0),
    ])
    def test_parse_amount(self, text, expected):
        assert core.parse_amount(text) == expected

    def test_amounts_sort_numerically_not_lexically(self):
        """字串排序會把 9,000 排在 21,000,000 之後，數值排序不會。"""
        values = ["9,000 元", "21,000,000 元", "100,000 元"]
        assert sorted(values, key=core.parse_amount) == [
            "9,000 元", "100,000 元", "21,000,000 元",
        ]


# ==================== 決標方式判定 ====================

class TestDetermineAwardMethod:
    @pytest.mark.parametrize("tender_way, expected_desc, expected_lowest", [
        ("公開招標", "最低標 (公開招標)", True),
        ("選擇性招標", "最低標 (選擇性招標)", True),
        ("公開取得報價單或企劃書", "公開取得 (待確認)", True),
        ("限制性招標", "限制性招標", False),
        ("公開評選", "最有利標 / 評選", False),
        ("", "未標明", False),
    ])
    def test_estimates_from_tender_way(self, tender_way, expected_desc, expected_lowest):
        assert core.determine_award_method(tender_way) == (expected_desc, expected_lowest)

    @pytest.mark.parametrize("actual, expected_lowest", [
        ("最低標", True),
        ("最低標，且不訂底價", True),
        ("最有利標", False),
        ("準用最有利標", False),
        ("參考最有利標", False),
        ("非以價格為考量之最有利標", False),
    ])
    def test_detail_page_value_wins(self, actual, expected_lowest):
        """詳細頁欄位為最高準則，且原文須原樣保留。"""
        desc, is_lowest = core.determine_award_method("公開招標", actual)
        assert desc == actual
        assert is_lowest is expected_lowest

    def test_reference_best_value_is_not_lowest_bid(self):
        """「參考最有利標」同時含最有利標字樣，不可被誤判為最低標。"""
        _desc, is_lowest = core.determine_award_method("公開招標", "參考最有利標")
        assert is_lowest is False

    def test_detail_page_overrides_optimistic_guess(self):
        """『公開取得』推估為最低標，但詳細頁若為評選則必須被推翻。"""
        _desc, guess = core.determine_award_method("公開取得報價單或企劃書")
        assert guess is True
        _desc, corrected = core.determine_award_method("公開取得報價單或企劃書", "最有利標")
        assert corrected is False

    def test_unknown_detail_value_is_not_lowest(self):
        assert core.determine_award_method("公開招標", "其他") == ("其他", False)


# ==================== 表頭欄位對照 ====================

class TestColumnIndex:
    def test_derives_indices_from_real_header(self, search_html):
        index_map, warnings = core.parse_column_index(search_html)
        assert warnings == []
        assert index_map == {
            "org": 1, "id_name": 2, "way": 4,
            "cate": 5, "pub": 6, "deadline": 7, "budget": 8,
        }

    def test_falls_back_and_warns_when_header_missing(self):
        """網站改版導致表頭消失時，必須發出警告而非靜默錯位。"""
        index_map, warnings = core.parse_column_index("<table><tr><td>x</td></tr></table>")
        assert index_map == core.DEFAULT_COLUMN_INDEX
        assert warnings and "表頭" in warnings[0]

    def test_warns_on_renamed_column(self):
        html = ("<table><tr><th>項次</th><th>機關名稱</th><th>標案案號標案名稱</th>"
                "<th>傳輸次數</th><th>招標方式</th><th>採購類別</th>"
                "<th>公告日期</th><th>截止投標</th><th>預算金額</th></tr></table>")
        index_map, warnings = core.parse_column_index(html)
        assert any("採購性質" in w for w in warnings)
        assert index_map["cate"] == core.DEFAULT_COLUMN_INDEX["cate"]

    def test_tracks_shifted_columns(self):
        """若網站在前面插入一欄，索引必須跟著位移。"""
        html = ("<table><tr><th>核選</th><th>項次</th><th>機關名稱</th>"
                "<th>標案案號標案名稱</th><th>傳輸次數</th><th>招標方式</th>"
                "<th>採購性質</th><th>公告日期</th><th>截止投標</th>"
                "<th>預算金額</th></tr></table>")
        index_map, warnings = core.parse_column_index(html)
        assert warnings == []
        assert index_map["org"] == 2
        assert index_map["budget"] == 9


# ==================== 搜尋結果解析 ====================

class TestParseTenderRows:
    def test_parses_expected_row_count(self, search_html):
        assert len(core.parse_tender_rows(search_html, "AI")) == 7

    def test_parses_full_page(self, paged_html):
        assert len(core.parse_tender_rows(paged_html, "系統")) == 50

    def test_first_row_fields(self, search_html):
        row = core.parse_tender_rows(search_html, "AI")[0]
        assert row["標案案號"] == "115JZ037"
        assert row["標案名稱"] == "115年「AI軟硬體暨教育訓練服務採購案」"
        assert row["招標機關"] == "交通部鐵道局"
        assert row["招標方式"] == "公開招標"
        assert row["採購性質"] == "勞務類"
        assert row["預算金額"] == "2,608,500 元"
        assert row["搜尋關鍵字"] == "AI"
        assert row["詳細連結"].endswith(row["pk"])

    def test_both_date_columns_are_western(self, search_html):
        """公告日期與截止投標必須同為西元，否則同表兩欄格式不一致。"""
        for row in core.parse_tender_rows(search_html, "AI"):
            assert row["公告日期"].startswith("20")
            assert row["截止投標"].startswith("20")

    def test_dates_sort_chronologically_as_strings(self, search_html):
        rows = core.parse_tender_rows(search_html, "AI")
        dates = [r["公告日期"] for r in rows]
        assert sorted(dates) == sorted(dates, key=lambda d: [int(p) for p in d.split("/")])

    def test_correction_notice_split_out_of_tender_id(self, search_html):
        """『(更正公告)』註記須移出標案案號，否則同一標案會被當成兩筆。"""
        rows = core.parse_tender_rows(search_html, "AI")
        corrected = [r for r in rows if r["公告類型"]]
        assert corrected, "fixture 應含至少一筆更正公告"
        assert corrected[0]["標案案號"] == "11508121"
        assert corrected[0]["公告類型"] == "更正公告"
        assert all("更正公告" not in r["標案案號"] for r in rows)

    def test_tender_name_never_falls_back_to_transfer_count(self):
        """
        迴歸測試：標案名稱與案號同屬第 2 欄，第 3 欄是「傳輸次數」。
        名稱取不到時必須退回同一格的連結文字，不可誤取傳輸次數。
        """
        html = """
        <table>
          <tr><th>項次</th><th>機關名稱</th><th>標案案號標案名稱</th><th>傳輸次數</th>
              <th>招標方式</th><th>採購性質</th><th>公告日期</th><th>截止投標</th>
              <th>預算金額</th></tr>
          <tr>
            <td>1</td><td>某某機關</td>
            <td>A123<br><a href="/prkms/urlSelector/common/tpam?pk=ABC123">純文字標案名稱</a></td>
            <td>07</td><td>公開招標</td><td>勞務類</td>
            <td>115/08/24</td><td>115/09/03</td><td>1,000</td>
          </tr>
        </table>
        """
        row = core.parse_tender_rows(html, "測試")[0]
        assert row["標案名稱"] == "純文字標案名稱"
        assert row["標案名稱"] != "07"

    def test_ignores_rows_without_pk(self):
        html = "<table><tr><td>1</td><td>2</td><td>3</td></tr></table>"
        assert core.parse_tender_rows(html, "kw") == []

    def test_flags_service_and_lowest_bid(self, search_html):
        for row in core.parse_tender_rows(search_html, "AI"):
            expected = "是" if "勞務" in row["採購性質"] else "否"
            assert row["是否為勞務類"] == expected


# ==================== 分頁 ====================

class TestPagination:
    def test_reads_total_records(self, paged_html):
        assert core.parse_total_records(paged_html) == 138

    def test_computes_total_pages(self, paged_html):
        assert core.parse_total_pages(paged_html) == 3

    def test_single_page_result_has_no_banner(self, search_html):
        assert core.parse_total_pages(search_html) == 1

    def test_extracts_displaytag_page_param(self, paged_html):
        assert core.parse_page_param(paged_html) == "d-49738-p"

    def test_page_param_also_discoverable_from_sort_links(self, search_html):
        """
        單頁結果不會渲染分頁列，但欄位排序連結同樣帶有該參數，
        因此參數名稱仍取得到（總頁數為 1 時不會被使用）。
        """
        assert core.parse_page_param(search_html) == "d-49738-p"

    def test_page_param_missing_when_markup_changes(self):
        assert core.parse_page_param("<table><tr><td>無分頁</td></tr></table>") is None

    def test_falls_back_to_first_page_when_param_missing(self, monkeypatch, paged_html):
        """找不到分頁參數時只回第 1 頁並示警，不可猜參數名亂送。"""
        html = paged_html.replace("d-49738-p", "d-XXXXX-q")
        monkeypatch.setattr(core, "http_post", lambda url, data, **kw: html)
        messages = []
        rows = core.search_pcc("系統", "115/06/01", "115/08/25",
                               log=messages.append, polite_delay=0)
        assert len(rows) == 50
        assert any("找不到分頁參數" in m for m in messages)

    @pytest.mark.parametrize("total, expected_pages", [
        (0, 1), (1, 1), (50, 1), (51, 2), (100, 2), (138, 3),
    ])
    def test_page_count_arithmetic(self, total, expected_pages):
        html = f'共有<span class="red"> {total} </span>筆資料'
        assert core.parse_total_pages(html) == expected_pages


class TestSearchPcc:
    """search_pcc 的翻頁行為（以假的 http_post 取代網路連線）。"""

    def test_walks_every_page(self, monkeypatch, paged_html, search_html):
        requested = []

        def fake_post(url, data, **kwargs):
            requested.append(dict(data))
            page = data.get("d-49738-p", "1")
            return paged_html if page == "1" else search_html

        monkeypatch.setattr(core, "http_post", fake_post)
        rows = core.search_pcc("系統", "115/06/01", "115/08/25", polite_delay=0)

        assert [r.get("d-49738-p") for r in requested] == [None, "2", "3"]
        # 第 1 頁 50 筆 + 第 2、3 頁各回 7 筆的替身內容
        assert len(rows) == 50 + 7 + 7

    def test_stops_at_max_pages(self, monkeypatch, paged_html, search_html):
        monkeypatch.setattr(core, "http_post",
                            lambda url, data, **kw: paged_html if "d-49738-p" not in data else search_html)
        rows = core.search_pcc("系統", "115/06/01", "115/08/25", max_pages=2, polite_delay=0)
        assert len(rows) == 50 + 7

    def test_single_page_does_not_paginate(self, monkeypatch, search_html):
        calls = []

        def fake_post(url, data, **kwargs):
            calls.append(data)
            return search_html

        monkeypatch.setattr(core, "http_post", fake_post)
        rows = core.search_pcc("AI", "115/08/01", "115/08/25", polite_delay=0)
        assert len(calls) == 1
        assert len(rows) == 7

    def test_captcha_page_returns_empty(self, monkeypatch):
        monkeypatch.setattr(core, "http_post", lambda url, data, **kw: "請輸入圖形驗證碼")
        monkeypatch.setattr(core.time, "sleep", lambda _s: None)
        assert core.search_pcc("AI", "115/08/01", "115/08/25", polite_delay=0) == []

    def test_connection_failure_is_reported_not_raised(self, monkeypatch):
        def boom(url, data, **kwargs):
            raise OSError("連線逾時")

        monkeypatch.setattr(core, "http_post", boom)
        messages = []
        assert core.search_pcc("AI", "115/08/01", "115/08/25",
                               log=messages.append, polite_delay=0) == []
        assert any("搜尋連線失敗" in m for m in messages)

    def test_proctrg_cate_is_forwarded(self, monkeypatch, search_html):
        captured = {}

        def fake_post(url, data, **kwargs):
            captured.update(data)
            return search_html

        monkeypatch.setattr(core, "http_post", fake_post)
        core.search_pcc("AI", "115/08/01", "115/08/25",
                        proctrg_cate=core.PROCTRG_CATE["工程"], polite_delay=0)
        assert captured["radProctrgCate"] == "RAD_PROCTRG_CATE_1"

    def test_unlimited_category_omits_the_parameter(self, monkeypatch, search_html):
        captured = {}

        def fake_post(url, data, **kwargs):
            captured.update(data)
            return search_html

        monkeypatch.setattr(core, "http_post", fake_post)
        core.search_pcc("AI", "115/08/01", "115/08/25", proctrg_cate=None, polite_delay=0)
        assert "radProctrgCate" not in captured


# ==================== 詳細頁校驗 ====================

class TestDetailPage:
    def test_extracts_award_method(self, monkeypatch, detail_html):
        monkeypatch.setattr(core, "http_get", lambda url, **kw: detail_html)
        assert core.fetch_actual_award_method("NzEzMDgzNTk=") == "最低標"

    def test_empty_pk_skips_request(self, monkeypatch):
        monkeypatch.setattr(core, "http_get",
                            lambda url, **kw: pytest.fail("不應對空 pk 發出請求"))
        assert core.fetch_actual_award_method("") == ""

    def test_network_error_degrades_quietly(self, monkeypatch):
        def boom(url, **kwargs):
            raise OSError("timeout")

        monkeypatch.setattr(core, "http_get", boom)
        assert core.fetch_actual_award_method("abc") == ""

    def test_apply_award_method_updates_derived_fields(self):
        tender = {
            "招標方式": "公開取得報價單或企劃書",
            "決標方式": "公開取得 (待確認)",
            "決標方式來源": core.AWARD_SOURCE_ESTIMATED,
            "是否為勞務類": "是",
            "是否為最低標": "是",
            "完全符合目標": "符合 (勞務+最低標)",
        }
        core.apply_award_method(tender, "最有利標")
        assert tender["決標方式"] == "最有利標"
        assert tender["是否為最低標"] == "否"
        assert tender["完全符合目標"] == "其他"
        assert tender["決標方式來源"] == core.AWARD_SOURCE_OFFICIAL

    def test_award_source_marks_unverified_rows(self, monkeypatch, search_html):
        """
        詳細頁取不到值時，該列必須留在「推估」狀態，
        使用者才看得出哪些「最低標」還沒經官方欄位確認。
        """
        rows = core.parse_tender_rows(search_html, "AI")
        assert all(r["決標方式來源"] == core.AWARD_SOURCE_ESTIMATED for r in rows)

        monkeypatch.setattr(core, "DETAIL_DELAY_RANGE", (0, 0))
        monkeypatch.setattr(core, "fetch_award_method_status",
                            lambda pk: ("最低標", "ok") if pk == rows[0]["pk"] else ("", "error"))
        core.enrich_actual_award_methods(rows)

        assert rows[0]["決標方式來源"] == core.AWARD_SOURCE_OFFICIAL
        assert all(r["決標方式來源"] == core.AWARD_SOURCE_ESTIMATED for r in rows[1:])

    def test_enrich_reports_failures(self, monkeypatch):
        monkeypatch.setattr(core, "DETAIL_DELAY_RANGE", (0, 0))
        monkeypatch.setattr(core, "fetch_award_method_status",
                            lambda pk: ("", "error") if pk == "b" else ("最低標", "ok"))
        tenders = [{"pk": "a", "招標方式": "公開招標", "是否為勞務類": "是"},
                   {"pk": "b", "招標方式": "公開招標", "是否為勞務類": "是"}]
        messages = []
        stats = core.enrich_actual_award_methods(tenders, log=messages.append)
        assert stats == {"total": 2, "done": 2, "ok": 1, "blocked": False,
                         "mirror_ok": 0, "official_ok": 1}
        assert any("1 筆" in m and "沒有決標方式欄位" in m for m in messages)

    def test_enrich_distinguishes_blocked_from_missing(self, monkeypatch):
        """
        「被限流擋下」稍後重試就拿得到，「公告沒這個欄位」重試幾次都一樣。
        兩者混為一談會讓使用者不知道還該不該再試。
        """
        monkeypatch.setattr(core, "DETAIL_DELAY_RANGE", (0, 0))
        monkeypatch.setattr(core, "fetch_award_method_status",
                            lambda pk: ("", "blocked") if pk == "a" else ("", "error"))
        tenders = [{"pk": "a", "招標方式": "公開招標", "是否為勞務類": "是"},
                   {"pk": "b", "招標方式": "公開招標", "是否為勞務類": "是"}]
        messages = []
        core.enrich_actual_award_methods(tenders, log=messages.append)

        assert any("限流擋下" in m and "1 筆" in m for m in messages)
        assert any("沒有決標方式欄位" in m and "1 筆" in m for m in messages)

    def test_enrich_stops_when_blocked_by_captcha(self, monkeypatch):
        """
        站方連續回「驗證碼檢核」頁時要立刻收手並明講，
        不能一路撞牆後讓整批資料靜默停在推估值，也不該為此拖長每次搜尋。
        """
        monkeypatch.setattr(core, "DETAIL_DELAY_RANGE", (0, 0))
        monkeypatch.setattr(core, "fetch_award_method_status", lambda pk: ("", "blocked"))
        tenders = [{"pk": str(i), "招標方式": "公開招標", "是否為勞務類": "是",
                    "決標方式來源": core.AWARD_SOURCE_ESTIMATED}
                   for i in range(30)]
        messages = []
        stats = core.enrich_actual_award_methods(tenders, log=messages.append)

        assert stats["blocked"] is True
        assert stats["ok"] == 0
        # done 計的是【真的送出去的請求】，不是走過的筆數；否則收手後訊息會說
        # 「剩餘 0 筆維持推估」，與實際上還有 25 筆沒查完完全相反。
        assert stats["done"] == core.CAPTCHA_STREAK_LIMIT
        assert all(t["決標方式來源"] == core.AWARD_SOURCE_ESTIMATED for t in tenders)
        assert any("額度已用盡" in m for m in messages)
        assert any(f"剩餘 {len(tenders) - core.CAPTCHA_STREAK_LIMIT} 筆" in m
                   for m in messages)

    def test_enrich_makes_no_further_requests_after_quota_runs_out(self, monkeypatch):
        """
        額度用盡後不該再打任何一次請求——繼續撞牆只會拖慢搜尋，
        而且每次重試都可能讓站方的 IP 冷卻重新計時。
        """
        monkeypatch.setattr(core, "DETAIL_DELAY_RANGE", (0, 0))
        calls = []

        def _fake(pk):
            calls.append(pk)
            return "", "blocked"

        monkeypatch.setattr(core, "fetch_award_method_status", _fake)
        tenders = [{"pk": str(i), "招標方式": "公開招標", "是否為勞務類": "是"}
                   for i in range(30)]
        core.enrich_actual_award_methods(tenders)

        assert len(calls) == core.CAPTCHA_STREAK_LIMIT

    def test_enrich_resets_streak_on_success(self, monkeypatch):
        """零星被擋不算數，只有連續被擋才代表真的被防護鎖住。"""
        monkeypatch.setattr(core, "DETAIL_DELAY_RANGE", (0, 0))
        monkeypatch.setattr(
            core, "fetch_award_method_status",
            lambda pk: ("", "blocked") if int(pk) % 2 else ("最低標", "ok"))
        tenders = [{"pk": str(i), "招標方式": "公開招標", "是否為勞務類": "是"}
                   for i in range(20)]
        stats = core.enrich_actual_award_methods(tenders)
        assert stats["blocked"] is False
        assert stats["ok"] == 10

    def test_enrich_reports_progress_for_every_item(self, monkeypatch):
        monkeypatch.setattr(core, "DETAIL_DELAY_RANGE", (0, 0))
        monkeypatch.setattr(core, "fetch_award_method_status", lambda pk: ("最低標", "ok"))
        tenders = [{"pk": str(i), "招標方式": "公開招標", "是否為勞務類": "是"} for i in range(10)]
        seen = []
        core.enrich_actual_award_methods(tenders,
                                         progress_cb=lambda d, t: seen.append((d, t)))
        assert sorted(d for d, _ in seen) == list(range(1, 11))
        assert all(t == 10 for _, t in seen)


# ==================== 去重、篩選與匯出 ====================

class TestMergeAndFilter:
    def test_merge_accumulates_keywords(self):
        unique = {}
        core.merge_by_tender_id(unique, [{"標案案號": "A1", "標案名稱": "案一"}], "AI")
        core.merge_by_tender_id(unique, [{"標案案號": "A1", "標案名稱": "案一"}], "資訊")
        core.merge_by_tender_id(unique, [{"標案案號": "A1", "標案名稱": "案一"}], "AI")
        assert list(unique) == ["A1"]
        assert unique["A1"]["命中關鍵字群"] == ["AI", "資訊"]

    def test_finalize_keywords_flattens(self):
        tenders = [{"命中關鍵字群": ["AI", "資訊"]}, {}]
        core.finalize_keywords(tenders)
        assert tenders[0]["命中關鍵字"] == "AI, 資訊"
        assert tenders[1]["命中關鍵字"] == ""

    @pytest.mark.parametrize("attr, award, expected_ids", [
        ("勞務", "最低標", ["s-low"]),
        ("勞務", "不限", ["s-low", "s-best"]),
        ("不限", "最低標", ["s-low", "g-low"]),
        ("不限", "不限", ["s-low", "s-best", "g-low"]),
        ("勞務", "最有利標/評選", ["s-best"]),
    ])
    def test_filter_combinations(self, attr, award, expected_ids):
        tenders = [
            {"標案案號": "s-low", "採購性質": "勞務類", "是否為最低標": "是", "決標方式": "最低標"},
            {"標案案號": "s-best", "採購性質": "勞務類", "是否為最低標": "否", "決標方式": "最有利標"},
            {"標案案號": "g-low", "採購性質": "財物類", "是否為最低標": "是", "決標方式": "最低標"},
        ]
        result = core.filter_tenders(tenders, attr, award)
        assert [t["標案案號"] for t in result] == expected_ids


class TestReports:
    def test_csv_excludes_internal_keys(self, tmp_path, search_html):
        rows = core.parse_tender_rows(search_html, "AI")
        core.finalize_keywords(rows)
        path = str(tmp_path / "out.csv")
        core.write_csv_report(path, rows)

        with open(path, encoding="utf-8-sig") as f:
            header = f.readline().strip().split(",")
        assert "pk" not in header
        assert "命中關鍵字群" not in header
        assert "標案名稱" in header

    def test_csv_of_empty_list_writes_nothing(self, tmp_path):
        path = str(tmp_path / "empty.csv")
        assert core.write_csv_report(path, []) == ""
        assert not (tmp_path / "empty.csv").exists()

    def test_excel_has_both_sheets_even_when_no_match(self, tmp_path, search_html):
        pytest.importorskip("openpyxl")
        rows = core.parse_tender_rows(search_html, "AI")
        core.finalize_keywords(rows)
        path = str(tmp_path / "out.xlsx")
        core.write_excel_report(path, rows, [])

        from openpyxl import load_workbook
        wb = load_workbook(path)
        assert wb.sheetnames == ["精選_勞務最低標", "所有搜尋標案"]
        assert wb["所有搜尋標案"].max_row == len(rows) + 1

    def test_excel_columns_follow_preferred_order(self, tmp_path, search_html):
        pytest.importorskip("openpyxl")
        rows = core.parse_tender_rows(search_html, "AI")
        core.finalize_keywords(rows)
        path = str(tmp_path / "out.xlsx")
        core.write_excel_report(path, rows, rows[:1])

        from openpyxl import load_workbook
        wb = load_workbook(path)
        header = [c.value for c in wb["所有搜尋標案"][1]]
        assert header == core.report_columns(rows)
        # PREFERRED_COLS 的欄位要照順序排在最前面
        assert header[:len(core.PREFERRED_COLS)] == core.PREFERRED_COLS

    def test_csv_and_excel_share_the_same_columns(self, tmp_path, search_html):
        """
        迴歸測試：同一批資料的兩種輸出曾經欄序不同（Excel 走 PREFERRED_COLS，
        CSV 走 rows[0].keys()），且 CSV 還多出三欄。使用者對照兩份檔案時會錯亂。
        """
        pytest.importorskip("openpyxl")
        rows = core.parse_tender_rows(search_html, "AI")
        core.finalize_keywords(rows)

        csv_path = str(tmp_path / "out.csv")
        xlsx_path = str(tmp_path / "out.xlsx")
        core.write_csv_report(csv_path, rows)
        core.write_excel_report(xlsx_path, rows, rows[:1])

        with open(csv_path, encoding="utf-8-sig") as f:
            csv_header = f.readline().strip().split(",")
        from openpyxl import load_workbook
        xlsx_header = [c.value for c in load_workbook(xlsx_path)["所有搜尋標案"][1]]

        assert csv_header == xlsx_header

    def test_csv_keeps_columns_that_only_later_rows_have(self, tmp_path):
        """
        先前以 rows[0].keys() 當欄位表且 extrasaction="ignore"，
        第一列沒有的鍵會被靜默丟掉——資料無聲消失比報錯更糟。
        """
        rows = [{"標案名稱": "甲案", "招標機關": "機關甲"},
                {"標案名稱": "乙案", "招標機關": "機關乙", "備註": "後來才出現的欄位"}]
        path = str(tmp_path / "out.csv")
        core.write_csv_report(path, rows)

        with open(path, encoding="utf-8-sig") as f:
            lines = [line.strip() for line in f]
        assert "備註" in lines[0].split(",")
        assert "後來才出現的欄位" in lines[2]
        # 缺該欄的列要留空，不能整列錯位
        assert lines[1].endswith(",")


# ==================== HTML 工具 ====================

class TestStripTags:
    @pytest.mark.parametrize("raw, expected", [
        ("<td>  文字  </td>", "文字"),
        ("<td>a<script>var x=1;</script>b</td>", "ab"),
        ("<td>甲&emsp;&emsp;乙</td>", "甲 乙"),
        ("<td>A&amp;B</td>", "A&B"),
        ("", ""),
    ])
    def test_strip_tags(self, raw, expected):
        assert core.strip_tags(raw) == expected


class TestCaptchaDetection:
    @pytest.mark.parametrize("html, expected", [
        ("請輸入圖形驗證碼", True),
        ("<p>請輸入驗證碼</p>", True),
        ("<div>驗證碼檢核</div>", True),
        ('<form id="validateForm" action="/tps/validate/check"></form>', True),
        ("<table>正常結果</table>", False),
    ])
    def test_is_captcha_page(self, html, expected):
        assert core.is_captcha_page(html) is expected

    def test_detail_page_captcha_is_recognised(self, captcha_html):
        """詳細頁的防護頁沒有「圖形驗證碼」字樣，仍必須被認出來。"""
        assert "圖形驗證碼" not in captcha_html
        assert core.is_captcha_page(captcha_html) is True

    def test_blocked_detail_page_reports_blocked_status(self, monkeypatch, captcha_html):
        monkeypatch.setattr(core, "http_get", lambda url, **kw: captcha_html)
        assert core.fetch_award_method_status("PK") == ("", "blocked")
        assert core.fetch_actual_award_method("PK") == ""

    def test_detail_page_status_ok(self, monkeypatch, detail_html):
        monkeypatch.setattr(core, "http_get", lambda url, **kw: detail_html)
        assert core.fetch_award_method_status("PK") == ("最低標", "ok")


# ==================== 查詢參數：日期模式與全面掃描 ====================

class TestSearchForm:
    """
    站方的兩種日期模式行為完全不同：
      isSpdt（等標期內）會忽略日期區間；isDate（公告日期區間）只吃西元日期，
      送民國日期一律回 0 筆。這裡把這兩件事釘住。
    """

    def test_range_mode_sends_ad_dates(self):
        form = core.build_search_form("AI", "115/08/01", "115/08/25",
                                      date_type=core.DATE_TYPE_RANGE)
        assert form["dateType"] == "isDate"
        assert form["tenderStartDate"] == "2026/08/01"
        assert form["tenderEndDate"] == "2026/08/25"

    def test_ad_dates_pass_through_unchanged(self):
        form = core.build_search_form("AI", "2026/08/01", "2026/08/25",
                                      date_type=core.DATE_TYPE_RANGE)
        assert form["tenderStartDate"] == "2026/08/01"

    def test_default_mode_is_bidding_period(self):
        form = core.build_search_form("AI", "2026/08/01", "2026/08/25")
        assert form["dateType"] == core.DATE_TYPE_SPDT

    def test_empty_keyword_scans_everything(self):
        form = core.build_search_form("", "2026/08/01", "2026/08/25",
                                      core.PROCTRG_CATE["勞務"])
        assert form["tenderName"] == ""
        assert form["radProctrgCate"] == "RAD_PROCTRG_CATE_3"

    @pytest.mark.parametrize("days, expected", [(7, 7), (186, 186), (365, core.MAX_RANGE_DAYS)])
    def test_range_days_are_clamped(self, days, expected):
        messages = []
        assert core.clamp_date_range_days(days, log=messages.append) == expected
        assert bool(messages) is (expected != days)


class TestFullScan:
    def test_date_type_is_forwarded(self, monkeypatch, search_html):
        captured = {}

        def fake_post(url, data, **kwargs):
            captured.update(data)
            return search_html

        monkeypatch.setattr(core, "http_post", fake_post)
        core.search_pcc("", "2026/08/01", "2026/08/25",
                        date_type=core.DATE_TYPE_RANGE, polite_delay=0)
        assert captured["dateType"] == "isDate"
        assert captured["tenderName"] == ""

    def test_truncation_is_reported_not_silent(self, monkeypatch, paged_html, search_html):
        """頁數超過上限時必須明講，不能像以前那樣靜默截斷。"""
        monkeypatch.setattr(core, "http_post",
                            lambda url, data, **kw: paged_html if "d-49738-p" not in data else search_html)
        messages = []
        core.search_pcc("", "2026/08/01", "2026/08/25", max_pages=2,
                        log=messages.append, polite_delay=0)
        assert any("頁數超過上限" in m and "未取回" in m for m in messages)

    def test_no_truncation_warning_within_limit(self, monkeypatch, paged_html, search_html):
        monkeypatch.setattr(core, "http_post",
                            lambda url, data, **kw: paged_html if "d-49738-p" not in data else search_html)
        messages = []
        core.search_pcc("", "2026/08/01", "2026/08/25", log=messages.append, polite_delay=0)
        assert not any("頁數超過上限" in m for m in messages)

    def test_page_progress_is_reported(self, monkeypatch, paged_html, search_html):
        monkeypatch.setattr(core, "http_post",
                            lambda url, data, **kw: paged_html if "d-49738-p" not in data else search_html)
        seen = []
        core.search_pcc("", "2026/08/01", "2026/08/25", polite_delay=0,
                        progress_cb=lambda done, total: seen.append((done, total)))
        assert seen == [(1, 3), (2, 3), (3, 3)]


class TestKeywordTagging:
    """關鍵字只標記、不篩選——這正是本次漏抓案例要守住的行為。"""

    MISSED = {"標案名稱": "桃園醫院人力資源E指通計畫採購案"}

    def test_tags_hit_keywords(self):
        rows = [dict(self.MISSED)]
        core.tag_keywords(rows, ["人力資源", "AI"])
        assert rows[0]["命中關鍵字群"] == ["人力資源"]

    def test_keeps_rows_without_any_hit(self):
        rows = [dict(self.MISSED)]
        core.tag_keywords(rows, ["AI", "資訊", "系統"])
        assert rows[0]["命中關鍵字群"] == []
        core.finalize_keywords(rows)
        assert rows[0]["命中關鍵字"] == ""

    def test_matching_is_case_insensitive(self):
        rows = [{"標案名稱": "校園 app 維運服務案"}]
        core.tag_keywords(rows, ["APP"])
        assert rows[0]["命中關鍵字群"] == ["APP"]

    def test_blank_keywords_are_ignored(self):
        rows = [dict(self.MISSED)]
        core.tag_keywords(rows, ["", "   ", "人力資源"])
        assert rows[0]["命中關鍵字群"] == ["人力資源"]


class TestEnrichmentSelection:
    """只校驗有機會入選的標案，才不會用上千次請求去撞驗證碼防護。"""

    @staticmethod
    def _row(pk, way, cate, pub):
        rows = core.parse_tender_rows("", "")  # 保持與正式流程相同的欄位語意
        assert rows == []
        desc, is_lowest = core.determine_award_method(way)
        return {
            "pk": pk, "標案案號": pk, "招標方式": way, "採購性質": cate, "公告日期": pub,
            "決標方式": desc, "決標方式來源": core.AWARD_SOURCE_ESTIMATED,
            "是否為勞務類": "是" if "勞務" in cate else "否",
            "是否為最低標": "是" if is_lowest else "否",
        }

    def test_skips_rows_that_cannot_qualify(self):
        rows = [
            self._row("a", "公開招標", "勞務類", "2026/08/20"),
            self._row("b", "經公開評選或公開徵求之限制性招標", "勞務類", "2026/08/21"),
            self._row("c", "公開招標", "工程類", "2026/08/22"),
        ]
        picked = core.select_rows_for_enrichment(rows, "勞務", "最低標")
        assert [r["pk"] for r in picked] == ["a"]

    def test_newest_first_and_limited(self):
        rows = [self._row(str(i), "公開招標", "勞務類", f"2026/08/{i:02d}") for i in range(1, 11)]
        picked = core.select_rows_for_enrichment(rows, "勞務", "最低標", limit=3)
        assert [r["公告日期"] for r in picked] == ["2026/08/10", "2026/08/09", "2026/08/08"]

    def test_zero_limit_means_unlimited(self):
        rows = [self._row(str(i), "公開招標", "勞務類", "2026/08/20") for i in range(5)]
        assert len(core.select_rows_for_enrichment(rows, "勞務", "最低標", limit=0)) == 5

    def test_selection_shares_the_same_row_objects(self):
        """挑出來的是同一批 dict，校驗結果才會回寫到原始清單。"""
        rows = [self._row("a", "公開招標", "勞務類", "2026/08/20")]
        picked = core.select_rows_for_enrichment(rows, "勞務", "最低標")
        assert picked[0] is rows[0]

    def test_skips_rows_already_confirmed(self):
        """
        額度稀缺（站方每輪只給約 5 筆詳細頁），已由快取確認過的標案不該再花一次額度。
        """
        rows = [
            self._row("a", "公開取得報價單或企劃書", "勞務類", "2026/08/20"),
            self._row("b", "公開取得報價單或企劃書", "勞務類", "2026/08/21"),
        ]
        core.apply_award_method(rows[1], "最低標")
        picked = core.select_rows_for_enrichment(rows, "勞務", "最低標")
        assert [r["pk"] for r in picked] == ["a"]


class TestAwardCache:
    """
    站方每輪只給約 5 筆詳細頁額度，一次執行不可能校驗完整批標案。
    確認過的結果必須落地累積，否則每次重跑都從零開始，永遠補不完。
    """

    @staticmethod
    def _row(tender_id, way="公開取得報價單或企劃書"):
        desc, is_lowest = core.determine_award_method(way)
        return {
            "pk": f"pk-{tender_id}", "標案案號": tender_id, "招標方式": way,
            "採購性質": "勞務類", "決標方式": desc,
            "決標方式來源": core.AWARD_SOURCE_ESTIMATED,
            "是否為勞務類": "是", "是否為最低標": "是" if is_lowest else "否",
        }

    def test_round_trip(self, tmp_path):
        path = core.award_cache_path(str(tmp_path))
        cache = {}
        core.remember_award(cache, self._row("A1"), "最低標")
        core.save_award_cache(cache, path)

        loaded = core.load_award_cache(path)
        assert loaded["A1"]["決標方式"] == "最低標"
        assert loaded["A1"]["pk"] == "pk-A1"
        assert loaded["A1"]["verified_at"]

    def test_missing_file_returns_empty(self, tmp_path):
        assert core.load_award_cache(str(tmp_path / "nope.json")) == {}

    def test_corrupt_file_returns_empty(self, tmp_path):
        path = tmp_path / "award_cache.json"
        path.write_text("{ not json", encoding="utf-8")
        assert core.load_award_cache(str(path)) == {}

    def test_non_dict_payload_returns_empty(self, tmp_path):
        path = tmp_path / "award_cache.json"
        path.write_text("[1, 2, 3]", encoding="utf-8")
        assert core.load_award_cache(str(path)) == {}

    def test_apply_cache_updates_derived_fields(self):
        """
        套快取必須連帶更新衍生欄位，否則精選清單仍會用舊的「是否為最低標」收案。
        """
        rows = [self._row("A1"), self._row("A2")]
        cache = {"A1": {"決標方式": "參考最有利標精神"}}

        assert core.apply_award_cache(rows, cache) == 1
        assert rows[0]["決標方式"] == "參考最有利標精神"
        assert rows[0]["是否為最低標"] == "否"
        assert rows[0]["完全符合目標"] == "其他"
        assert core.is_award_confirmed(rows[0])
        # 沒被快取涵蓋的那筆維持推估，不能被誤標成已確認
        assert rows[1]["決標方式"] == "公開取得 (待確認)"
        assert not core.is_award_confirmed(rows[1])

    def test_apply_cache_removes_them_from_enrichment(self):
        """套過快取的標案不會再被挑去校驗，額度才花得到刀口上。"""
        rows = [self._row("A1"), self._row("A2")]
        core.apply_award_cache(rows, {"A1": {"決標方式": "最低標"}})
        picked = core.select_rows_for_enrichment(rows, "勞務", "最低標")
        assert [r["標案案號"] for r in picked] == ["A2"]

    def test_empty_cache_is_a_noop(self):
        rows = [self._row("A1")]
        assert core.apply_award_cache(rows, {}) == 0
        assert not core.is_award_confirmed(rows[0])

    def test_enrich_persists_each_success_immediately(self, monkeypatch, tmp_path):
        """
        被擋而中止時，已經花掉額度換來的結果必須已經落地，
        否則下次重跑又要重新花額度確認同一批標案。
        """
        monkeypatch.setattr(core, "DETAIL_DELAY_RANGE", (0, 0))
        # 前兩筆成功，之後一路被擋直到中止
        monkeypatch.setattr(
            core, "fetch_award_method_status",
            lambda pk: ("最低標", "ok") if pk in ("pk-A0", "pk-A1") else ("", "blocked"))

        rows = [self._row(f"A{i}") for i in range(30)]
        path = core.award_cache_path(str(tmp_path))
        cache = {}
        stats = core.enrich_actual_award_methods(rows,
                                                 cache=cache, cache_path=path)

        assert stats["blocked"] is True
        assert sorted(core.load_award_cache(path)) == ["A0", "A1"]

    def test_enrich_without_cache_still_works(self, monkeypatch):
        """未傳快取時行為不變，CLI 或測試才能單獨呼叫。"""
        monkeypatch.setattr(core, "DETAIL_DELAY_RANGE", (0, 0))
        monkeypatch.setattr(core, "fetch_award_method_status", lambda pk: ("最低標", "ok"))
        rows = [self._row("A1")]
        assert core.enrich_actual_award_methods(rows)["ok"] == 1


class TestCancellation:
    """
    全面掃描動輒 120 頁要跑好幾分鐘，使用者發現條件設錯時必須能中斷，
    而且已經抓到的資料沒有理由丟掉。
    """

    def test_search_stops_paging_but_keeps_partial_results(self, monkeypatch, search_html,
                                                           paged_html):
        stop = {"now": False}
        pages = []

        def _fake_post(url, form, **kwargs):
            pages.append(form.get("d-3611-p"))
            if len(pages) >= 2:
                stop["now"] = True          # 抓完第 2 頁後使用者按下停止
            return paged_html

        monkeypatch.setattr(core, "http_post", _fake_post)
        monkeypatch.setattr(core, "PAGE_DELAY", 0)
        messages = []
        rows = core.search_pcc("", "2026/08/01", "2026/08/24", polite_delay=0,
                               log=messages.append, should_stop=lambda: stop["now"])

        assert len(pages) == 2, "停止後不該再翻下一頁"
        assert rows, "已抓到的資料必須保留，不能因為中斷就整批丟掉"
        assert any("已取消翻頁" in m for m in messages)

    def test_search_without_callback_is_unaffected(self, monkeypatch, paged_html):
        """未傳 should_stop 時行為不變（CLI 走這條路）。"""
        monkeypatch.setattr(core, "http_post", lambda *a, **kw: paged_html)
        monkeypatch.setattr(core, "PAGE_DELAY", 0)
        rows = core.search_pcc("", "2026/08/01", "2026/08/24", polite_delay=0, max_pages=3)
        assert len(rows) == 150

    def test_enrich_stops_on_request(self, monkeypatch):
        stop = {"now": False}
        calls = []

        def _fake(pk):
            calls.append(pk)
            if len(calls) >= 3:
                stop["now"] = True
            return "最低標", "ok"

        monkeypatch.setattr(core, "DETAIL_DELAY_RANGE", (0, 0))
        monkeypatch.setattr(core, "fetch_award_method_status", _fake)
        tenders = [{"pk": str(i), "招標方式": "公開招標", "是否為勞務類": "是"}
                   for i in range(20)]
        stats = core.enrich_actual_award_methods(tenders,
                                                 should_stop=lambda: stop["now"])

        assert len(calls) == 3
        assert stats["ok"] == 3
        # 停止前確認到的必須保留
        assert all(core.is_award_confirmed(t) for t in tenders[:3])


class TestKeywordHitFiltering:
    """
    全面掃描會撈回整批勞務標案（午餐、粉刷、校外教學…），
    精選清單要靠「命中關鍵字」這層才有意義；但未命中者只是不進精選，不會被丟掉。
    """

    @staticmethod
    def _row(name, hits, cate="勞務類", lowest="是"):
        return {"標案名稱": name, "命中關鍵字群": list(hits), "採購性質": cate,
                "是否為最低標": lowest, "決標方式": "最低標", "公告日期": "2026/08/20"}

    def test_default_keeps_rows_without_hits(self):
        rows = [self._row("內湖區樓梯間粉刷案", [])]
        assert len(core.filter_tenders(rows, "勞務", "最低標")) == 1

    def test_require_hit_drops_rows_without_hits(self):
        rows = [self._row("內湖區樓梯間粉刷案", []),
                self._row("桃園醫院人力資源E指通計畫採購案", ["人力資源", "E指通"])]
        picked = core.filter_tenders(rows, "勞務", "最低標", require_keyword_hit=True)
        assert [r["標案名稱"] for r in picked] == ["桃園醫院人力資源E指通計畫採購案"]

    def test_flattened_keyword_string_also_counts(self):
        """finalize_keywords() 之後只剩字串欄位，一樣要認得。"""
        row = self._row("智慧路燈維護案", ["維護"])
        core.finalize_keywords([row])
        del row["命中關鍵字群"]
        assert core.has_keyword_hit(row) is True
        assert len(core.filter_tenders([row], "勞務", "最低標", require_keyword_hit=True)) == 1

    def test_hit_alone_is_not_enough(self):
        """命中關鍵字但決標方式不符時，仍不該進精選。"""
        rows = [self._row("市府資訊系統評選案", ["資訊", "系統"], lowest="否")]
        assert core.filter_tenders(rows, "勞務", "最低標", require_keyword_hit=True) == []

    def test_enrichment_selection_follows_the_same_rule(self):
        rows = [self._row("內湖區樓梯間粉刷案", []),
                self._row("智慧路燈維護案", ["維護"])]
        for row in rows:
            row["pk"] = row["標案名稱"]
            row["招標方式"] = "公開招標"
        picked = core.select_rows_for_enrichment(rows, "勞務", "最低標", require_keyword_hit=True)
        assert [r["標案名稱"] for r in picked] == ["智慧路燈維護案"]


# ==================== 巢狀表格 ====================

class TestNestedTables:
    """
    先前以 `<tr[^>]*>(.*?)</tr>` 切列，遇到儲存格內還有一層 table 時，
    外層列會被【內層的 </tr> 提早切斷】，後面的欄位整批錯位或整列被丟掉。
    """

    NESTED = ("<table><tr><td>甲</td>"
              "<td><table><tr><td>內層</td></tr></table></td>"
              "<td>丙</td></tr></table>")

    def test_outer_row_keeps_every_cell(self):
        rows = core.iter_table_rows(self.NESTED)
        outer = [r for r in rows if r["depth"] == 1]
        assert len(outer) == 1
        assert len(outer[0]["cells"]) == 3, "外層列不該被內層的 </tr> 切斷"
        assert core.strip_tags(outer[0]["cells"][2]) == "丙"

    def test_inner_row_is_reported_separately(self):
        inner = [r for r in core.iter_table_rows(self.NESTED) if r["depth"] == 2]
        assert [core.strip_tags(c) for c in inner[0]["cells"]] == ["內層"]

    def test_rows_follow_document_order(self):
        html = "<table><tr><td>一</td></tr><tr><td>二</td></tr><tr><td>三</td></tr></table>"
        rows = core.iter_table_rows(html)
        assert [core.strip_tags(r["cells"][0]) for r in rows] == ["一", "二", "三"]

    def test_header_and_data_cells_are_distinguished(self):
        html = "<table><tr><th>標題</th></tr><tr><td>內容</td></tr></table>"
        rows = core.iter_table_rows(html)
        assert rows[0]["tags"] == ["th"]
        assert rows[1]["tags"] == ["td"]

    def test_missing_close_tags_do_not_break_parsing(self):
        """站方漏掉 </td> / </tr> 時要同層自動收尾，不能整份解析崩掉。"""
        html = "<table><tr><td>甲<td>乙<tr><td>丙</table>"
        rows = core.iter_table_rows(html)
        assert [core.strip_tags(c) for c in rows[0]["cells"]] == ["甲", "乙"]
        assert [core.strip_tags(c) for c in rows[1]["cells"]] == ["丙"]

    def test_real_rows_still_parse_after_wrapping_in_a_nested_table(self, search_html):
        """把真實搜尋結果整個包進另一層 table，解析結果必須完全一樣。"""
        plain = core.parse_tender_rows(search_html, "AI")
        wrapped = core.parse_tender_rows(
            "<table><tr><td>" + search_html + "</td></tr></table>", "AI")
        assert [r["標案案號"] for r in wrapped] == [r["標案案號"] for r in plain]


# ==================== 驗證碼的結構性判斷 ====================

class TestCaptchaStructuralDetection:
    """
    字串比對是站方換文案就靜默失效的那種偵測：我們會把驗證碼頁當成「查無資料」，
    使用者只看到 0 筆而不知道自己被擋了。表單結構是比較不會變的後盾。
    """

    def test_detects_validate_form_with_unknown_wording(self):
        html = ("<html><body><h1>Security Check</h1>"
                "<form id=\"f\" action=\"/tps/validate/check\" method=\"post\"></form>"
                "</body></html>")
        assert core.is_captcha_page(html) is True

    def test_detects_validate_code_input(self):
        html = "<html><form><input type=\"text\" name=\"validateCode\" /></form></html>"
        assert core.is_captcha_page(html) is True

    def test_normal_search_page_is_not_flagged(self, search_html):
        assert core.is_captcha_page(search_html) is False

    def test_normal_detail_page_is_not_flagged(self, detail_html):
        assert core.is_captcha_page(detail_html) is False


# ==================== 快取有效期 ====================

class TestAwardCacheTTL:
    """
    更正公告會改動決標方式。「一次確認、永久相信」遲早會拿舊答案餵給使用者，
    而且比「待確認」更危險——後者至少畫成橘色提醒使用者自己去看。
    """

    @staticmethod
    def _row(tender_id="A1"):
        return {"pk": "pk-" + tender_id, "標案案號": tender_id,
                "招標方式": "公開取得報價單或企劃書", "採購性質": "勞務類",
                "決標方式": "公開取得 (待確認)",
                "決標方式來源": core.AWARD_SOURCE_ESTIMATED,
                "是否為勞務類": "是", "是否為最低標": "是"}

    @staticmethod
    def _entry(days_ago, award="最低標"):
        stamp = date.today() - timedelta(days=days_ago)
        return {"決標方式": award, "pk": "pk-A1", "verified_at": stamp.isoformat()}

    def test_fresh_entry_is_applied(self):
        rows = [self._row()]
        assert core.apply_award_cache(rows, {"A1": self._entry(1)}) == 1
        assert core.is_award_confirmed(rows[0])

    def test_expired_entry_is_ignored(self):
        rows = [self._row()]
        stale = {"A1": self._entry(core.AWARD_CACHE_TTL_DAYS + 1)}
        assert core.apply_award_cache(rows, stale) == 0
        assert not core.is_award_confirmed(rows[0])
        assert rows[0]["決標方式"] == "公開取得 (待確認)"

    def test_entry_exactly_on_the_boundary_still_counts(self):
        rows = [self._row()]
        entry = {"A1": self._entry(core.AWARD_CACHE_TTL_DAYS)}
        assert core.apply_award_cache(rows, entry) == 1

    def test_legacy_entries_without_timestamp_are_kept(self):
        """早期版本寫的純字串條目沒有時間戳，不該因為升級就整批失效。"""
        rows = [self._row()]
        assert core.apply_award_cache(rows, {"A1": "最低標"}) == 1

    def test_unparseable_timestamp_is_treated_as_fresh(self):
        rows = [self._row()]
        entry = {"決標方式": "最低標", "verified_at": "not-a-date"}
        assert core.apply_award_cache(rows, {"A1": entry}) == 1

    def test_expired_rows_go_back_into_the_verification_queue(self):
        """過期就該重新排隊去查，否則永遠拿著舊答案。"""
        rows = [self._row()]
        core.apply_award_cache(rows, {"A1": self._entry(core.AWARD_CACHE_TTL_DAYS + 1)})
        picked = core.select_rows_for_enrichment(rows, "勞務", "最低標")
        assert [r["標案案號"] for r in picked] == ["A1"]

    def test_prune_drops_only_expired_entries(self):
        cache = {"A1": self._entry(1), "A2": self._entry(core.AWARD_CACHE_TTL_DAYS + 5)}
        assert core.prune_award_cache(cache) == 1
        assert list(cache) == ["A1"]


# ==================== 待確認佇列 ====================

class TestPendingQueue:
    """背景涓流校驗靠這個檔案知道「還有哪些要查」，不必為了拿清單重跑整個全掃。"""

    @staticmethod
    def _row(tender_id):
        return {"pk": "pk-" + tender_id, "標案案號": tender_id,
                "標案名稱": tender_id + " 案", "招標機關": "某機關",
                "招標方式": "公開取得報價單或企劃書"}

    def test_round_trip(self, tmp_path):
        path = core.pending_queue_path(str(tmp_path))
        assert core.save_pending_queue([self._row("A1"), self._row("A2")], path) is True
        rows = core.load_pending_queue(path)
        assert sorted(r["標案案號"] for r in rows) == ["A1", "A2"]
        assert all(r["pk"].startswith("pk-") for r in rows)

    def test_rows_without_pk_are_skipped(self, tmp_path):
        """沒有 pk 就連不了詳細頁，留在佇列裡只會每輪浪費一次額度。"""
        path = core.pending_queue_path(str(tmp_path))
        bad = self._row("A9")
        del bad["pk"]
        core.save_pending_queue([self._row("A1"), bad], path)
        assert [r["標案案號"] for r in core.load_pending_queue(path)] == ["A1"]

    def test_already_cached_rows_are_filtered_out(self, tmp_path):
        path = core.pending_queue_path(str(tmp_path))
        core.save_pending_queue([self._row("A1"), self._row("A2")], path)
        cache = {"A1": {"決標方式": "最低標", "verified_at": date.today().isoformat()}}
        assert [r["標案案號"] for r in core.load_pending_queue(path, cache)] == ["A2"]

    def test_expired_cache_entries_go_back_into_the_queue(self, tmp_path):
        path = core.pending_queue_path(str(tmp_path))
        core.save_pending_queue([self._row("A1")], path)
        stale = date.today() - timedelta(days=core.AWARD_CACHE_TTL_DAYS + 1)
        cache = {"A1": {"決標方式": "最低標", "verified_at": stale.isoformat()}}
        assert [r["標案案號"] for r in core.load_pending_queue(path, cache)] == ["A1"]

    def test_priority_order_survives_the_round_trip(self, tmp_path):
        """
        JSON 以案號排序存放，不另外記順序的話取件會變成字母序（A0, A1, A10, A11, A2…），
        額度就先花在早已過了等標期的舊案上。呼叫端給的順序才是優先順序。
        """
        path = core.pending_queue_path(str(tmp_path))
        core.save_pending_queue([self._row("A10"), self._row("A2"), self._row("A1")], path)
        assert [r["標案案號"] for r in core.load_pending_queue(path)] == ["A10", "A2", "A1"]

    def test_missing_file_is_an_empty_queue(self, tmp_path):
        assert core.load_pending_queue(str(tmp_path / "nope.json")) == []

    def test_corrupt_file_is_an_empty_queue(self, tmp_path):
        path = tmp_path / "pending_queue.json"
        path.write_text("{ not json", encoding="utf-8")
        assert core.load_pending_queue(str(path)) == []


# ==================== 背景涓流校驗 ====================

class TestTrickleVerify:
    """
    補完待確認清單靠的是每輪撿一批、跨輪累積。
    每一輪都要能獨立跑完並把成果落地，中途被擋也不能把佇列弄丟。
    """

    @staticmethod
    def _seed(tmp_path, count=12):
        queue_path = core.pending_queue_path(str(tmp_path))
        rows = [{"pk": "pk-A%d" % i, "標案案號": "A%d" % i, "標案名稱": "第 %d 案" % i,
                 "招標機關": "某機關", "招標方式": "公開取得報價單或企劃書"}
                for i in range(count)]
        core.save_pending_queue(rows, queue_path)
        return core.award_cache_path(str(tmp_path)), queue_path

    def test_one_round_only_takes_a_small_batch(self, monkeypatch, tmp_path):
        monkeypatch.setattr(core, "DETAIL_DELAY_RANGE", (0, 0))
        calls = []

        def _fake(pk):
            calls.append(pk)
            return "最低標", "ok"

        monkeypatch.setattr(core, "fetch_award_method_status", _fake)
        cache_path, queue_path = self._seed(tmp_path)
        result = core.trickle_verify(cache_path, queue_path, batch=5)

        assert len(calls) == 5, "一輪就該收手，不要硬啃整份清單"
        assert result["ok"] == 5

    def test_default_batch_matches_the_mirror_throughput(self, monkeypatch, tmp_path):
        """
        預設批量跟著主來源走：鏡像沒有官網那種 5 筆/輪的額度，
        還把每輪壓在 5 筆，待確認清單只會愈積愈多。
        """
        monkeypatch.setattr(core, "DETAIL_DELAY_RANGE", (0, 0))
        monkeypatch.setattr(core, "fetch_award_method_status", lambda pk: ("最低標", "ok"))
        cache_path, queue_path = self._seed(tmp_path, count=core.DEFAULT_TRICKLE_BATCH + 7)

        result = core.trickle_verify(cache_path, queue_path)

        assert result["picked"] == core.DEFAULT_TRICKLE_BATCH
        assert result["remaining"] == 7

    def test_confirmed_rows_leave_the_queue_and_land_in_the_cache(self, monkeypatch, tmp_path):
        monkeypatch.setattr(core, "DETAIL_DELAY_RANGE", (0, 0))
        monkeypatch.setattr(core, "fetch_award_method_status", lambda pk: ("最低標", "ok"))
        cache_path, queue_path = self._seed(tmp_path, count=12)

        result = core.trickle_verify(cache_path, queue_path, batch=5)

        assert len(core.load_award_cache(cache_path)) == 5
        assert result["remaining"] == 12 - 5
        assert len(core.load_pending_queue(queue_path)) == 12 - 5

    def test_rounds_accumulate_across_calls(self, monkeypatch, tmp_path):
        """這是整個設計的重點：跨次執行慢慢累積，而不是每次歸零。"""
        monkeypatch.setattr(core, "DETAIL_DELAY_RANGE", (0, 0))
        monkeypatch.setattr(core, "fetch_award_method_status", lambda pk: ("最低標", "ok"))
        cache_path, queue_path = self._seed(tmp_path, count=12)

        confirmed = []
        for _ in range(3):
            confirmed.append(core.trickle_verify(cache_path, queue_path, batch=5)["ok"])

        # 12 筆分三輪撿完：5 + 5 + 2，最後一輪只剩 2 筆可撿
        assert confirmed == [5, 5, 2]
        assert len(core.load_award_cache(cache_path)) == 12
        assert len(core.load_pending_queue(queue_path)) == 0

    def test_blocked_round_keeps_the_queue_intact(self, monkeypatch, tmp_path):
        """被擋下時什麼都沒確認到，佇列就該原封不動留給下一輪。"""
        monkeypatch.setattr(core, "DETAIL_DELAY_RANGE", (0, 0))
        monkeypatch.setattr(core, "fetch_award_method_status", lambda pk: ("", "blocked"))
        cache_path, queue_path = self._seed(tmp_path, count=12)

        result = core.trickle_verify(cache_path, queue_path)

        assert result["blocked"] is True
        assert result["ok"] == 0
        assert len(core.load_pending_queue(queue_path)) == 12

    def test_empty_queue_makes_no_requests(self, monkeypatch, tmp_path):
        def _boom(pk):
            raise AssertionError("佇列是空的就不該連線")

        monkeypatch.setattr(core, "fetch_award_method_status", _boom)
        result = core.trickle_verify(core.award_cache_path(str(tmp_path)),
                                     core.pending_queue_path(str(tmp_path)))
        assert result == {"picked": 0, "ok": 0, "blocked": False, "remaining": 0}

    def test_stop_callback_is_honoured(self, monkeypatch, tmp_path):
        monkeypatch.setattr(core, "DETAIL_DELAY_RANGE", (0, 0))
        monkeypatch.setattr(core, "fetch_award_method_status", lambda pk: ("最低標", "ok"))
        cache_path, queue_path = self._seed(tmp_path)
        result = core.trickle_verify(cache_path, queue_path, should_stop=lambda: True)
        assert result["ok"] == 0


# ==================== 寫檔失敗的可見性 ====================

class TestWriteFailuresAreReported:
    """
    決標方式的累積策略成敗全繫於這個寫入。先前這裡靜默吞掉 OSError，
    磁碟滿了或資料夾唯讀時，使用者會以為一切正常，實際上什麼都沒存下來。
    """

    def test_successful_write_returns_true(self, tmp_path):
        assert core.save_json_dict({"a": 1}, str(tmp_path / "x.json")) is True

    def test_failed_write_returns_false(self, monkeypatch, tmp_path):
        def _boom(*args, **kwargs):
            raise OSError("disk full")

        monkeypatch.setattr(core.os, "replace", _boom)
        assert core.save_json_dict({"a": 1}, str(tmp_path / "x.json")) is False

    def test_failed_cache_write_is_logged(self, monkeypatch, tmp_path):
        monkeypatch.setattr(core, "DETAIL_DELAY_RANGE", (0, 0))
        monkeypatch.setattr(core, "fetch_award_method_status", lambda pk: ("最低標", "ok"))
        monkeypatch.setattr(core, "save_award_cache", lambda cache, path: False)
        messages = []
        rows = [{"pk": "pk-A1", "標案案號": "A1", "招標方式": "公開招標",
                 "是否為勞務類": "是"}]
        core.enrich_actual_award_methods(rows, log=messages.append, cache={},
                                         cache_path=str(tmp_path / "c.json"))
        assert any("無法寫入決標方式快取" in m for m in messages)

    def test_original_file_survives_a_failed_write(self, tmp_path):
        """先寫暫存檔再 os.replace：寫壞了也不該毀掉上一次的成果。"""
        path = str(tmp_path / "cache.json")
        core.save_json_dict({"good": 1}, path)
        try:
            core.save_json_dict({"bad": object()}, path)  # 不可序列化
        except TypeError:
            pass
        assert core.load_json_dict(path) == {"good": 1}
